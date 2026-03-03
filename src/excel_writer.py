import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path
from datetime import datetime


# ── Global style constants ───────────────────────────────────────────────────
DEFAULT_FONT_NAME = "Trebuchet MS"
DEFAULT_FONT_SIZE = 11

DEFAULT_FONT        = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE)
BOLD_FONT           = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True)
BOLD_WHITE_FONT     = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, bold=True, color="FFFFFF")
GREY_FONT           = Font(name=DEFAULT_FONT_NAME, size=DEFAULT_FONT_SIZE, color="AAAAAA")

CENTER_ALIGN        = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN          = Alignment(horizontal="left",   vertical="center", wrap_text=True)

HEADER_FILL         = PatternFill("solid", fgColor="1F4E79")
SUBHEADER_FILL      = PatternFill("solid", fgColor="2E75B6")
RULE_ID_FILL        = PatternFill("solid", fgColor="D6E4F0")
GREEN_FILL          = PatternFill("solid", fgColor="C6EFCE")
AMBER_FILL          = PatternFill("solid", fgColor="FFEB9C")
RED_FILL            = PatternFill("solid", fgColor="FFC7CE")

# All-sides thin border applied to every cell
_THIN  = Side(style="thin", color="000000")
ALL_BORDERS = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style(cell,
           font=None,
           fill=None,
           alignment=None,
           border=True):
    """Apply font, fill, alignment and border to a cell in one call."""
    cell.font      = font      or DEFAULT_FONT
    cell.alignment = alignment or CENTER_ALIGN
    if fill:
        cell.fill  = fill
    if border:
        cell.border = ALL_BORDERS


def _score_fill(score: float) -> PatternFill:
    if score >= 0.8:
        return GREEN_FILL
    elif score >= 0.6:
        return AMBER_FILL
    return RED_FILL


def write_excel(results: list[dict], output_dir: str = None) -> str:
    """
    Write ranked retrieval results to a formatted Excel file.

    Sheet 1 — Summary  : Transposed. Rows = Rules, Columns = Vendors.
    Sheet 2 — Full Detail : Original flat layout, kept for reference.

    Returns the full path of the created file.
    """
    if output_dir is None:
        output_dir = Path(__file__).parent.parent / "output"

    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename  = f"SyDRe_Results_{timestamp}.xlsx"
    filepath  = output_path / filename

    wb = openpyxl.Workbook()

    # ════════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Summary (Rules × Vendors)
    # ════════════════════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Summary"

    # --- Collect ordered unique vendors and rules ---
    vendors      = []
    seen_vendors = set()
    rules        = []
    seen_rules   = set()

    for r in results:
        v   = r.get("vendor_id", "")
        rid = r.get("rule_id",   "")
        if v   and v   not in seen_vendors:
            vendors.append(v)
            seen_vendors.add(v)
        if rid and rid not in seen_rules:
            rules.append({"rule_id": rid, "rule_text": r.get("rule_text", "")})
            seen_rules.add(rid)

    rules.sort(key=lambda x: x["rule_id"])

    # --- Lookup: (rule_id, vendor_id) → best result (rank 1) ---
    lookup = {}
    for r in results:
        key = (r.get("rule_id", ""), r.get("vendor_id", ""))
        if key not in lookup or r.get("rank", 99) < lookup[key].get("rank", 99):
            lookup[key] = r

    # ── Row 1: Vendor merged headers ─────────────────────────────────────────
    for fixed_col, label in [(1, "Rule ID"), (2, "Rule Text")]:
        cell = ws.cell(row=1, column=fixed_col, value=label)
        _style(cell, font=BOLD_WHITE_FONT, fill=HEADER_FILL)

    for v_idx, vendor in enumerate(vendors):
        start_col = 3 + (v_idx * 3)
        end_col   = start_col + 2
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1,   end_column=end_col
        )
        cell = ws.cell(row=1, column=start_col, value=vendor)
        _style(cell, font=BOLD_WHITE_FONT, fill=HEADER_FILL)

        for mc in range(start_col, end_col + 1):
            ws.cell(row=1, column=mc).border = ALL_BORDERS

    ws.row_dimensions[1].height = 28

    # ── Row 2: Sub-headers (File Name | Page | Score) ────────────────────────
    for fixed_col, label in [(1, "Rule ID"), (2, "Rule Text")]:
        cell = ws.cell(row=2, column=fixed_col, value=label)
        _style(cell, font=BOLD_WHITE_FONT, fill=SUBHEADER_FILL)

    for v_idx in range(len(vendors)):
        for s_idx, sub in enumerate(["File Name", "Page", "Score"]):
            col  = 3 + (v_idx * 3) + s_idx
            cell = ws.cell(row=2, column=col, value=sub)
            _style(cell, font=BOLD_WHITE_FONT, fill=SUBHEADER_FILL)

    ws.row_dimensions[2].height = 25

    # ── Data rows: one row per rule ───────────────────────────────────────────
    for r_idx, rule in enumerate(rules):
        row = r_idx + 3

        cell = ws.cell(row=row, column=1, value=rule["rule_id"])
        _style(cell, font=BOLD_FONT, fill=RULE_ID_FILL)

        cell = ws.cell(row=row, column=2, value=rule["rule_text"])
        _style(cell, font=DEFAULT_FONT, fill=RULE_ID_FILL, alignment=LEFT_ALIGN)

        for v_idx, vendor in enumerate(vendors):
            key      = (rule["rule_id"], vendor)
            result   = lookup.get(key)
            base_col = 3 + (v_idx * 3)

            if result:
                score = result.get("score", 0)
                fill  = _score_fill(score)

                file_cell  = ws.cell(row=row, column=base_col,     value=result.get("file_name",   ""))
                page_cell  = ws.cell(row=row, column=base_col + 1, value=result.get("page_number", ""))
                score_cell = ws.cell(row=row, column=base_col + 2, value=score)

                for cell in (file_cell, page_cell, score_cell):
                    _style(cell, font=DEFAULT_FONT, fill=fill)
            else:
                for offset in range(3):
                    cell = ws.cell(row=row, column=base_col + offset, value="—")
                    _style(cell, font=GREY_FONT)

        ws.row_dimensions[row].height = 35

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 45

    for v_idx in range(len(vendors)):
        base = 3 + (v_idx * 3)
        ws.column_dimensions[openpyxl.utils.get_column_letter(base)    ].width = 28
        ws.column_dimensions[openpyxl.utils.get_column_letter(base + 1)].width = 8
        ws.column_dimensions[openpyxl.utils.get_column_letter(base + 2)].width = 10

    ws.freeze_panes = "C3"

    # ════════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Full Detail (original flat layout)
    # ════════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet(title="Full Detail")

    detail_headers = [
        "Rule ID", "Rule Text", "Rank", "Vendor ID",
        "File Name", "Page Number", "Similarity Score",
        "Source Type", "Text Snippet"
    ]

    for col_idx, header in enumerate(detail_headers, start=1):
        cell = ws2.cell(row=1, column=col_idx, value=header)
        _style(cell, font=BOLD_WHITE_FONT, fill=HEADER_FILL)

    ws2.row_dimensions[1].height = 30

    for row_idx, r in enumerate(results, start=2):
        score = r.get("score", 0)
        fill  = _score_fill(score)

        values = [
            r.get("rule_id",     ""),
            r.get("rule_text",   ""),
            r.get("rank",        ""),
            r.get("vendor_id",   ""),
            r.get("file_name",   ""),
            r.get("page_number", ""),
            score,
            r.get("source_type", ""),
            r.get("snippet",     ""),
        ]

        for col_idx, value in enumerate(values, start=1):
            align = LEFT_ALIGN if col_idx == 9 else CENTER_ALIGN
            cell  = ws2.cell(row=row_idx, column=col_idx, value=value)
            _style(cell, font=DEFAULT_FONT, fill=fill, alignment=align)

    detail_col_widths = [10, 50, 8, 15, 30, 14, 18, 14, 70]
    for col_idx, width in enumerate(detail_col_widths, start=1):
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    ws2.freeze_panes = "A2"

    wb.save(filepath)
    print(f"[excel_writer] Results saved to: {filepath}")
    return str(filepath)